
import os, re, datetime, random, string, smtplib, ssl
from tkinter import *
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from email.message import EmailMessage

ruta_excel = 'Ventas.xlsx'

color_primario = '#486f99'
color_secundario = '#5d83ae'
color_terciario = '#688db9'
color_resaltado = '#a6bddc'
color_fondo = '#e2e8f3'
color_claro = '#ffffff'
color_suave = '#c4d2e7'
color_boton = '#5379a4'
color_boton_2 = '#7398c4'
color_texto = '#0b213a'

def generar_id_ai(prefijo):
    t = datetime.datetime.utcnow().strftime('%y%m%d%H%M%S%f')
    suf = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
    return f'{prefijo}-{t[-10:]}{suf}'

def asegurar_excel():
    if not os.path.exists(ruta_excel):
        wb = Workbook()
        hs = wb.active
        hs.title = 'productos'
        hs.append(['id_ai','codigo','nombre','existencia','proveedor','precio'])
        hc = wb.create_sheet('clientes')
        hc.append(['id_ai','codigo','nombre','direccion','telefono','correo','fecha_registro'])
        hv = wb.create_sheet('ventas')
        hv.append(['id','codigo_producto','codigo_cliente','cantidad','total','fecha','estado'])
        wb.save(ruta_excel)
    else:
        wb = load_workbook(ruta_excel)
        if 'productos' in wb.sheetnames:
            ws = wb['productos']
            if (ws['A1'].value or '').lower() != 'id_ai':
                ws.insert_cols(1); ws['A1'] = 'id_ai'
                for i in range(2, ws.max_row+1):
                    if not ws.cell(row=i, column=1).value:
                        ws.cell(row=i, column=1).value = generar_id_ai('PR')
        else:
            ws = wb.create_sheet('productos')
            ws.append(['id_ai','codigo','nombre','existencia','proveedor','precio'])
        if 'clientes' in wb.sheetnames:
            ws = wb['clientes']
            if (ws['A1'].value or '').lower() != 'id_ai':
                ws.insert_cols(1); ws['A1'] = 'id_ai'
                for i in range(2, ws.max_row+1):
                    if not ws.cell(row=i, column=1).value:
                        ws.cell(row=i, column=1).value = generar_id_ai('CL')
        else:
            ws = wb.create_sheet('clientes')
            ws.append(['id_ai','codigo','nombre','direccion','telefono','correo','fecha_registro'])
        if 'ventas' not in wb.sheetnames:
            ws = wb.create_sheet('ventas')
            ws.append(['id','codigo_producto','codigo_cliente','cantidad','total','fecha','estado'])
        wb.save(ruta_excel)

def cargar_libro():
    asegurar_excel()
    return load_workbook(ruta_excel)

def guardar_libro(wb):
    wb.save(ruta_excel)

def obtener_hoja(nombre):
    wb = cargar_libro()
    if nombre not in wb.sheetnames:
        sh = wb.create_sheet(nombre)
    else:
        sh = wb[nombre]
    return wb, sh

def generar_codigo(tabla, prefijo):
    wb, sh = obtener_hoja(tabla)
    mayor = 0
    for i, fila in enumerate(sh.iter_rows(values_only=True)):
        if i == 0: continue
        if not fila or not any(fila): continue
        cod = str(fila[1] or '')
        m = re.match(rf'^{prefijo}-(\d+)$', cod)
        if m:
            try:
                mayor = max(mayor, int(m.group(1)))
            except:
                pass
    return f'{prefijo}-{mayor+1:05d}'

def listar_productos():
    wb, sh = obtener_hoja('productos')
    datos = []
    for i, fila in enumerate(sh.iter_rows(values_only=True)):
        if i == 0: continue
        if not fila or not any(fila): continue
        fila = list(fila) + [None]*6
        datos.append([fila[1], fila[2], fila[3], fila[4], fila[5]])
    return datos

def crear_producto(nombre, existencia, proveedor, precio):
    if not nombre:
        return False, 'El nombre es obligatorio'
    try:
        existencia = float(existencia)
        precio = float(precio)
    except:
        return False, 'Existencia y precio deben ser numéricos'
    if existencia < 0 or precio < 0:
        return False, 'Existencia y precio no pueden ser negativos'
    wb, sh = obtener_hoja('productos')
    codigo = generar_codigo('productos', 'PRC')
    id_ai = generar_id_ai('PR')
    sh.append([id_ai, codigo, nombre, float(existencia), proveedor, float(precio)])
    guardar_libro(wb)
    return True, 'Producto creado'

def actualizar_producto(codigo, nombre, existencia, proveedor, precio):
    try:
        existencia = float(existencia); precio = float(precio)
    except:
        return False, 'Existencia y precio deben ser numéricos'
    if existencia < 0 or precio < 0:
        return False, 'Existencia y precio no pueden ser negativos'
    wb, sh = obtener_hoja('productos')
    for idx, fila in enumerate(sh.iter_rows(values_only=True), start=1):
        if idx == 1: continue
        if fila and str(fila[1]) == str(codigo):
            sh.cell(row=idx, column=3).value = nombre
            sh.cell(row=idx, column=4).value = float(existencia)
            sh.cell(row=idx, column=5).value = proveedor
            sh.cell(row=idx, column=6).value = float(precio)
            guardar_libro(wb)
            return True, 'Producto actualizado'
    return False, 'No encontrado'

def editar_existencia(codigo, nueva_existencia):
    try: nueva_existencia = float(nueva_existencia)
    except: return False, 'Existencia debe ser numérica'
    if nueva_existencia < 0: return False, 'Existencia no puede ser negativa'
    wb, sh = obtener_hoja('productos')
    for idx, fila in enumerate(sh.iter_rows(values_only=True), start=1):
        if idx == 1: continue
        if fila and str(fila[1]) == str(codigo):
            sh.cell(row=idx, column=4).value = float(nueva_existencia)
            guardar_libro(wb); return True, 'Existencia actualizada'
    return False, 'No encontrado'

def eliminar_producto(codigo):
    wb, sh = obtener_hoja('productos')
    fila_borrar = None
    for idx, fila in enumerate(sh.iter_rows(values_only=True), start=1):
        if idx == 1: continue
        if fila and str(fila[1]) == str(codigo):
            fila_borrar = idx; break
    if not fila_borrar: return False, 'No encontrado'
    sh.delete_rows(fila_borrar, 1); guardar_libro(wb); return True, 'Eliminado'

def listar_clientes():
    wb, sh = obtener_hoja('clientes')
    datos = []
    for i, fila in enumerate(sh.iter_rows(values_only=True)):
        if i == 0: continue
        if not fila or not any(fila): continue
        fila = list(fila) + [None]*7
        datos.append([fila[1], fila[2], fila[3], fila[4], fila[5], fila[6]])
    return datos

def crear_cliente(nombre, direccion, telefono, correo):
    if not nombre:
        return False, 'El nombre es obligatorio'
    if correo and not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', correo):
        return False, 'Correo inválido'
    wb, sh = obtener_hoja('clientes')
    codigo = generar_codigo('clientes', 'CLC')
    id_ai = generar_id_ai('CL')
    fecha = datetime.datetime.now().strftime('%Y-%m-%d')
    sh.append([id_ai, codigo, nombre, direccion, telefono, correo, fecha])
    guardar_libro(wb)
    return True, 'Cliente creado'

def editar_cliente(codigo, nombre, direccion, telefono, correo):
    if correo and not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', correo):
        return False, 'Correo inválido'
    wb, sh = obtener_hoja('clientes')
    for idx, fila in enumerate(sh.iter_rows(values_only=True), start=1):
        if idx == 1: continue
        if fila and str(fila[1]) == str(codigo):
            sh.cell(row=idx, column=3).value = nombre
            sh.cell(row=idx, column=4).value = direccion
            sh.cell(row=idx, column=5).value = telefono
            sh.cell(row=idx, column=6).value = correo
            guardar_libro(wb); return True, 'Cliente actualizado'
    return False, 'No encontrado'

def eliminar_cliente(codigo):
    wb, sh = obtener_hoja('clientes')
    fila_borrar = None
    for idx, fila in enumerate(sh.iter_rows(values_only=True), start=1):
        if idx == 1: continue
        if fila and str(fila[1]) == str(codigo):
            fila_borrar = idx; break
    if not fila_borrar: return False, 'No encontrado'
    sh.delete_rows(fila_borrar, 1); guardar_libro(wb); return True, 'Eliminado'

def exportar_clientes(ruta_salida):
    datos = listar_clientes()
    wb_out = Workbook(); ws = wb_out.active; ws.title = 'clientes'
    ws.append(['codigo','nombre','direccion','telefono','correo','fecha_registro'])
    for f in datos: ws.append(f)
    for c in range(1,7): ws.column_dimensions[get_column_letter(c)].width = 24
    wb_out.save(ruta_salida)

def importar_clientes(ruta_entrada):
    wb_in = load_workbook(ruta_entrada); ws = wb_in.active
    encabezado = [c.value for c in ws[1]]; ind = {k:i for i,k in enumerate(encabezado)}
    req_almenos = ['nombre']
    for r in req_almenos:
        if r not in ind: return False, 'Encabezado requerido: nombre (otros campos opcionales)'
    wb, sh = obtener_hoja('clientes')
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not fila or not any(fila): continue
        codigo = None
        if 'codigo' in ind and fila[ind['codigo']]:
            codigo = str(fila[ind['codigo']])
        else:
            codigo = generar_codigo('clientes', 'CLC')
        nombre = str(fila[ind['nombre']])
        direccion = str(fila[ind.get('direccion', None)] or '')
        telefono = str(fila[ind.get('telefono', None)] or '')
        correo = str(fila[ind.get('correo', None)] or '')
        fecha = datetime.datetime.now().strftime('%Y-%m-%d')
        id_ai = generar_id_ai('CL')
        sh.append([id_ai, codigo, nombre, direccion, telefono, correo, fecha])
    guardar_libro(wb); return True, 'Importación completa'

def obtener_producto(codigo):
    wb, sh = obtener_hoja('productos')
    for idx, fila in enumerate(sh.iter_rows(values_only=True), start=1):
        if idx == 1: continue
        if fila and str(fila[1]) == str(codigo):
            return {'id_ai':fila[0],'codigo':fila[1],'nombre':fila[2],'existencia':fila[3],'proveedor':fila[4],'precio':fila[5]}
    return None

def obtener_cliente(codigo):
    wb, sh = obtener_hoja('clientes')
    for idx, fila in enumerate(sh.iter_rows(values_only=True), start=1):
        if idx == 1: continue
        if fila and str(fila[1]) == str(codigo):
            return {'id_ai':fila[0],'codigo':fila[1],'nombre':fila[2],'direccion':fila[3],'telefono':fila[4],'correo':fila[5],'fecha_registro':fila[6]}
    return None

def listar_ventas():
    wb, sh = obtener_hoja('ventas')
    datos = []
    for i, fila in enumerate(sh.iter_rows(values_only=True)):
        if i == 0: continue
        if not fila or not any(fila): continue
        datos.append(list(fila))
    return datos

def crear_venta(codigo_producto, codigo_cliente, cantidad):
    try: cantidad = float(cantidad)
    except: return False, 'La cantidad debe ser numérica'
    if cantidad <= 0: return False, 'La cantidad debe ser mayor que cero'
    prod = obtener_producto(codigo_producto); 
    if not prod: return False, 'Producto no existe'
    cli = obtener_cliente(codigo_cliente); 
    if not cli: return False, 'Cliente no existe'
    if float(prod['existencia']) < cantidad: return False, 'No hay existencia suficiente'
    total = float(prod['precio']) * cantidad
    wb, sh = obtener_hoja('ventas'); ultimo_id = 0
    for i, fila in enumerate(sh.iter_rows(values_only=True)):
        if i == 0: continue
        if fila and fila[0]:
            try: ultimo_id = max(ultimo_id, int(str(fila[0]).split('-')[-1]))
            except: pass
    nuevo_id = f'VE-{ultimo_id+1}'
    fecha = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sh.append([nuevo_id, codigo_producto, codigo_cliente, float(cantidad), round(total,2), fecha, 'activa'])
    wb_prod, hoja_prod = obtener_hoja('productos')
    for idx, fila in enumerate(hoja_prod.iter_rows(values_only=True), start=1):
        if idx == 1: continue
        if fila and str(fila[1]) == str(codigo_producto):
            hoja_prod.cell(row=idx, column=4).value = float(fila[3]) - cantidad; break
    guardar_libro(wb); guardar_libro(wb_prod); return True, 'Venta creada'

def anular_venta(id_venta):
    wb, sh = obtener_hoja('ventas'); fila_editar = None; datos = None
    for idx, fila in enumerate(sh.iter_rows(values_only=True), start=1):
        if idx == 1: continue
        if fila and str(fila[0]) == str(id_venta): fila_editar = idx; datos = fila; break
    if not fila_editar: return False, 'No encontrada'
    if sh.cell(row=fila_editar, column=7).value == 'anulada': return False, 'Ya anulada'
    sh.cell(row=fila_editar, column=7).value = 'anulada'
    wb_prod, hoja_prod = obtener_hoja('productos')
    for idx, fila in enumerate(hoja_prod.iter_rows(values_only=True), start=1):
        if idx == 1: continue
        if fila and str(fila[1]) == str(datos[1]): hoja_prod.cell(row=idx, column=4).value = float(fila[3]) + float(datos[3]); break
    guardar_libro(wb); guardar_libro(wb_prod); return True, 'Venta anulada'

def generar_reporte_por_cliente(ruta_salida):
    ventas = listar_ventas(); mapa = {}
    for v in ventas:
        if v[6] != 'activa': continue
        cli = obtener_cliente(v[2]); nombre = cli['nombre'] if cli else v[2]
        if nombre not in mapa: mapa[nombre] = {'cantidad':0,'total':0}
        mapa[nombre]['cantidad'] += float(v[3]); mapa[nombre]['total'] += float(v[4])
    wb_out = Workbook(); ws = wb_out.active; ws.title = 'ventas_por_cliente'
    ws.append(['cliente','cantidad','total'])
    for k, val in mapa.items(): ws.append([k, val['cantidad'], round(val['total'],2)])
    for c in range(1,4): ws.column_dimensions[get_column_letter(c)].width = 24
    wb_out.save(ruta_salida)

def generar_reporte_por_producto(ruta_salida):
    ventas = listar_ventas(); mapa = {}
    for v in ventas:
        if v[6] != 'activa': continue
        prod = obtener_producto(v[1]); nombre = prod['nombre'] if prod else v[1]
        if nombre not in mapa: mapa[nombre] = {'cantidad':0,'total':0}
        mapa[nombre]['cantidad'] += float(v[3]); mapa[nombre]['total'] += float(v[4])
    wb_out = Workbook(); ws = wb_out.active; ws.title = 'ventas_por_producto'
    ws.append(['producto','cantidad','total'])
    for k, val in mapa.items(): ws.append([k, val['cantidad'], round(val['total'],2)])
    for c in range(1,4): ws.column_dimensions[get_column_letter(c)].width = 24
    wb_out.save(ruta_salida)

class App(Tk):
    def __init__(self):
        super().__init__()
        self.title('Sistema de Ventas - Algoritmos')
        self.geometry('1100x700')
        self.configure(bg=color_fondo)
        self.resizable(False, False)
        self.estilo = ttk.Style(self)
        self.estilo.theme_use('clam')
        self.estilo.configure('TFrame', background=color_fondo)
        self.estilo.configure('Sub.TLabel', background=color_fondo, foreground=color_texto, font=('Segoe UI', 11, 'bold'))
        self.estilo.configure('TButton', padding=8)
        self.menu_lateral()
        self.cuerpo = ttk.Frame(self, padding=10, style='TFrame')
        self.cuerpo.place(x=220, y=10, width=870, height=680)
        self.vista_bienvenida()

    def menu_lateral(self):
        self.panel = ttk.Frame(self, padding=0)
        self.panel.place(x=10, y=10, width=200, height=680)
        cab = Frame(self.panel, bg=color_primario, height=70)
        cab.pack(fill='x')
        Label(cab, text='Menú Principal', bg=color_primario, fg=color_claro, font=('Segoe UI', 14, 'bold')).pack(padx=10, pady=20, anchor='w')
        for txt, cmd in [('Algoritmos', self.vista_algoritmos), ('Matemática Discreta', self.vista_md), ('Álgebra Lineal', self.vista_al)]:
            b = Button(self.panel, text=txt, bg=color_boton, fg=color_claro, activebackground=color_boton_2, font=('Segoe UI', 11, 'bold'), bd=0, relief='flat', height=2, command=cmd)
            b.pack(fill='x', pady=8)

    def limpiar_cuerpo(self):
        for w in self.cuerpo.winfo_children():
            w.destroy()

    def vista_bienvenida(self):
        self.limpiar_cuerpo()
        marco = ttk.Frame(self.cuerpo, padding=30)
        marco.pack(expand=True, fill='both')
        Label(marco, text='Proyecto Final: Sistemas de Ventas', bg=color_fondo, fg=color_texto, font=('Segoe UI', 22, 'bold')).pack(pady=10)
        Label(marco, text='Seleccione un módulo desde el menú', bg=color_fondo, fg=color_texto, font=('Segoe UI', 12)).pack(pady=10)

    def vista_md(self):
        self.limpiar_cuerpo()
        ttk.Label(self.cuerpo, text='Matemática Discreta (próximamente)', style='Sub.TLabel').pack(pady=20)

    def vista_al(self):
        self.limpiar_cuerpo()
        ttk.Label(self.cuerpo, text='Álgebra Lineal (próximamente)', style='Sub.TLabel').pack(pady=20)

    def vista_algoritmos(self):
        self.limpiar_cuerpo()
        cuaderno = ttk.Notebook(self.cuerpo)
        cuaderno.pack(expand=True, fill='both')
        self.tab_productos = ttk.Frame(cuaderno, padding=10)
        self.tab_clientes = ttk.Frame(cuaderno, padding=10)
        self.tab_ventas = ttk.Frame(cuaderno, padding=10)
        self.tab_reportes = ttk.Frame(cuaderno, padding=10)
        self.tab_correo = ttk.Frame(cuaderno, padding=10)
        cuaderno.add(self.tab_productos, text='Productos')
        cuaderno.add(self.tab_clientes, text='Clientes')
        cuaderno.add(self.tab_ventas, text='Ventas')
        cuaderno.add(self.tab_reportes, text='Reportes')
        cuaderno.add(self.tab_correo, text='Correo')
        self.ui_productos()
        self.ui_clientes()
        self.ui_ventas()
        self.ui_reportes()
        self.ui_correo()

    def ui_productos(self):
        sup = ttk.Frame(self.tab_productos); sup.pack(fill='x')
        Label(sup, text='Gestión de Productos', bg=color_fondo, fg=color_texto, font=('Segoe UI', 14, 'bold')).pack(side='left')
        barra = ttk.Frame(self.tab_productos); barra.pack(fill='x', pady=6)
        ttk.Button(barra, text='Refrescar', command=self.refrescar_productos).pack(side='left', padx=5)
        ttk.Button(barra, text='Nuevo', command=self.nuevo_producto).pack(side='left', padx=5)
        ttk.Button(barra, text='Editar', command=self.editar_producto).pack(side='left', padx=5)
        ttk.Button(barra, text='Existencia', command=self.cambiar_existencia).pack(side='left', padx=5)
        ttk.Button(barra, text='Eliminar', command=self.eliminar_producto).pack(side='left', padx=5)
        cont = ttk.Frame(self.tab_productos); cont.pack(expand=True, fill='both')
        cols = ('codigo','nombre','existencia','proveedor','precio')
        self.tabla_productos = ttk.Treeview(cont, columns=cols, show='headings', height=14)
        for c,t in zip(cols, ('Código','Nombre','Existencia','Proveedor','Precio')):
            self.tabla_productos.heading(c, text=t); self.tabla_productos.column(c, width=160 if c!='existencia' else 120, anchor='center')
        vsb = ttk.Scrollbar(cont, orient='vertical', command=self.tabla_productos.yview)
        self.tabla_productos.configure(yscrollcommand=vsb.set)
        self.tabla_productos.grid(row=0, column=0, sticky='nsew'); vsb.grid(row=0, column=1, sticky='ns')
        cont.rowconfigure(0, weight=1); cont.columnconfigure(0, weight=1)
        self.refrescar_productos()

    def ui_clientes(self):
        sup = ttk.Frame(self.tab_clientes); sup.pack(fill='x')
        Label(sup, text='Gestión de Clientes', bg=color_fondo, fg=color_texto, font=('Segoe UI', 14, 'bold')).pack(side='left')
        barra = ttk.Frame(self.tab_clientes); barra.pack(fill='x', pady=6)
        ttk.Button(barra, text='Refrescar', command=self.refrescar_clientes).pack(side='left', padx=5)
        ttk.Button(barra, text='Nuevo', command=self.nuevo_cliente).pack(side='left', padx=5)
        ttk.Button(barra, text='Editar', command=self.editar_cliente).pack(side='left', padx=5)
        ttk.Button(barra, text='Eliminar', command=self.eliminar_cliente).pack(side='left', padx=5)
        ttk.Button(barra, text='Exportar', command=self.exportar_clientes_ui).pack(side='left', padx=20)
        ttk.Button(barra, text='Importar', command=self.importar_clientes_ui).pack(side='left')
        busq = ttk.Frame(self.tab_clientes); busq.pack(fill='x', pady=4)
        ttk.Label(busq, text='Buscar:', style='Sub.TLabel').pack(side='left', padx=5)
        self.ent_buscar_cliente = ttk.Entry(busq, width=35); self.ent_buscar_cliente.pack(side='left', padx=5)
        self.ent_buscar_cliente.bind('<KeyRelease>', lambda e: self.filtrar_clientes())
        cont = ttk.Frame(self.tab_clientes); cont.pack(expand=True, fill='both')
        cols = ('codigo','nombre','direccion','telefono','correo','fecha_registro')
        self.tabla_clientes = ttk.Treeview(cont, columns=cols, show='headings', height=14)
        for c,t in zip(cols, ('Código','Nombre','Dirección','Teléfono','Correo','Fecha')):
            self.tabla_clientes.heading(c, text=t); self.tabla_clientes.column(c, width=150, anchor='center')
        vsb = ttk.Scrollbar(cont, orient='vertical', command=self.tabla_clientes.yview)
        self.tabla_clientes.configure(yscrollcommand=vsb.set)
        self.tabla_clientes.grid(row=0, column=0, sticky='nsew'); vsb.grid(row=0, column=1, sticky='ns')
        cont.rowconfigure(0, weight=1); cont.columnconfigure(0, weight=1)
        self.lbl_sin_clientes = ttk.Label(self.tab_clientes, text='No hay clientes. Usa \"Nuevo\" o \"Importar\".', style='Sub.TLabel')
        self.lbl_sin_clientes.pack_forget()
        self.refrescar_clientes()

    def ui_ventas(self):
        sup = ttk.Frame(self.tab_ventas); sup.pack(fill='x')
        ttk.Label(sup, text='Control de Ventas', style='Sub.TLabel').pack(side='left')
        form = ttk.Frame(self.tab_ventas, padding=10); form.pack(fill='x', pady=10)
        ttk.Label(form, text='Producto:').grid(row=0, column=0, sticky='w'); self.cb_producto = ttk.Combobox(form, state='readonly'); self.cb_producto.grid(row=0, column=1, sticky='ew', padx=5)
        ttk.Label(form, text='Cliente:').grid(row=0, column=2, sticky='w'); self.cb_cliente = ttk.Combobox(form, state='readonly'); self.cb_cliente.grid(row=0, column=3, sticky='ew', padx=5)
        ttk.Label(form, text='Cantidad:').grid(row=1, column=0, sticky='w'); self.ent_cantidad = ttk.Entry(form); self.ent_cantidad.grid(row=1, column=1, sticky='ew', padx=5)
        self.lbl_total = ttk.Label(form, text='Total: Q0.00'); self.lbl_total.grid(row=1, column=2, sticky='w')
        ttk.Button(form, text='Calcular', command=self.calcular_total).grid(row=1, column=3, padx=5)
        ttk.Button(form, text='Crear Venta', command=self.crear_venta_ui).grid(row=2, column=0, pady=8)
        ttk.Button(form, text='Anular Venta', command=self.anular_venta_ui).grid(row=2, column=1, pady=8)
        for i in range(4): form.columnconfigure(i, weight=1)
        barra = ttk.Frame(self.tab_ventas); barra.pack(fill='x', pady=4)
        ttk.Label(barra, text='Buscar:', style='Sub.TLabel').pack(side='left', padx=5)
        self.ent_buscar_venta = ttk.Entry(barra, width=40); self.ent_buscar_venta.pack(side='left', padx=5)
        self.ent_buscar_venta.bind('<KeyRelease>', lambda e: self.filtrar_ventas())
        cont = ttk.Frame(self.tab_ventas); cont.pack(expand=True, fill='both')
        cols = ('id','codigo_producto','codigo_cliente','cantidad','total','fecha','estado')
        self.tabla_ventas = ttk.Treeview(cont, columns=cols, show='headings', height=12)
        for c,t in zip(cols, ('ID','Producto','Cliente','Cantidad','Total','Fecha','Estado')):
            self.tabla_ventas.heading(c, text=t); self.tabla_ventas.column(c, width=120, anchor='center')
        vsb = ttk.Scrollbar(cont, orient='vertical', command=self.tabla_ventas.yview)
        self.tabla_ventas.configure(yscrollcommand=vsb.set)
        self.tabla_ventas.grid(row=0, column=0, sticky='nsew'); vsb.grid(row=0, column=1, sticky='ns')
        cont.rowconfigure(0, weight=1); cont.columnconfigure(0, weight=1)
        self.refrescar_combobox(); self.refrescar_ventas()


        # --- Panel inferior: Venta reciente ---
        ttk.Separator(self.tab_ventas, orient='horizontal').pack(fill='x', pady=6)
        self.panel_venta = ttk.Frame(self.tab_ventas, padding=10)
        self.panel_venta.pack(fill='x')

        ttk.Label(self.panel_venta, text='Venta reciente', style='Sub.TLabel').grid(row=0, column=0, columnspan=8, sticky='w', pady=(0,6))

        # Variables para mostrar detalle
        self.v_id = StringVar(); self.v_prod = StringVar(); self.v_cli = StringVar()
        self.v_cant = StringVar(); self.v_total = StringVar(); self.v_fecha = StringVar(); self.v_estado = StringVar()

        # Encabezados
        ttk.Label(self.panel_venta, text='ID:').grid(row=1, column=0, sticky='e', padx=(0,4))
        ttk.Label(self.panel_venta, text='Producto:').grid(row=1, column=2, sticky='e', padx=(12,4))
        ttk.Label(self.panel_venta, text='Cliente:').grid(row=1, column=4, sticky='e', padx=(12,4))

        ttk.Label(self.panel_venta, text='Cantidad:').grid(row=2, column=0, sticky='e', padx=(0,4), pady=(4,0))
        ttk.Label(self.panel_venta, text='Total:').grid(row=2, column=2, sticky='e', padx=(12,4), pady=(4,0))
        ttk.Label(self.panel_venta, text='Fecha:').grid(row=2, column=4, sticky='e', padx=(12,4), pady=(4,0))
        ttk.Label(self.panel_venta, text='Estado:').grid(row=2, column=6, sticky='e', padx=(12,4), pady=(4,0))

        # Valores
        ttk.Label(self.panel_venta, textvariable=self.v_id).grid(row=1, column=1, sticky='w')
        ttk.Label(self.panel_venta, textvariable=self.v_prod).grid(row=1, column=3, sticky='w')
        ttk.Label(self.panel_venta, textvariable=self.v_cli).grid(row=1, column=5, sticky='w')

        ttk.Label(self.panel_venta, textvariable=self.v_cant).grid(row=2, column=1, sticky='w', pady=(4,0))
        ttk.Label(self.panel_venta, textvariable=self.v_total).grid(row=2, column=3, sticky='w', pady=(4,0))
        ttk.Label(self.panel_venta, textvariable=self.v_fecha).grid(row=2, column=5, sticky='w', pady=(4,0))
        ttk.Label(self.panel_venta, textvariable=self.v_estado).grid(row=2, column=7, sticky='w', pady=(4,0))

        for i in range(8):
            self.panel_venta.columnconfigure(i, weight=1)

        # Iniciar con última venta si existe
            self.actualizar_detalle_venta_desde_ultima()
    def ui_reportes(self):
        cont = ttk.Frame(self.tab_reportes, padding=10); cont.pack(fill='x', pady=10)
        ttk.Button(cont, text='Ventas por Cliente (Excel)', command=self.reporte_cliente).pack(side='left', padx=5)
        ttk.Button(cont, text='Ventas por Producto (Excel)', command=self.reporte_producto).pack(side='left', padx=5)
        self.archivo_rep_cliente = StringVar(); self.archivo_rep_producto = StringVar()

    def ui_correo(self):
        cont = ttk.Frame(self.tab_correo, padding=10); cont.pack(fill='x', pady=10)
        ttk.Label(cont, text='Correo remitente (Gmail):').grid(row=0, column=0, sticky='w'); self.ent_correo = ttk.Entry(cont, width=40); self.ent_correo.grid(row=0, column=1)
        ttk.Label(cont, text='Contraseña de aplicación:').grid(row=1, column=0, sticky='w'); self.ent_pass = ttk.Entry(cont, width=40, show='*'); self.ent_pass.grid(row=1, column=1)
        ttk.Label(cont, text='Correo destino:').grid(row=2, column=0, sticky='w'); self.ent_destino = ttk.Entry(cont, width=40); self.ent_destino.grid(row=2, column=1)
        ttk.Button(cont, text='Enviar reportes adjuntos', command=self.enviar_reportes).grid(row=3, column=0, columnspan=2, pady=8)
        self.estado_correo = StringVar(); ttk.Label(cont, textvariable=self.estado_correo).grid(row=4, column=0, columnspan=2, sticky='w')

    # --- Productos (acciones)
    def refrescar_productos(self):
        if hasattr(self, 'tabla_productos'):
            for i in self.tabla_productos.get_children(): self.tabla_productos.delete(i)
            for f in listar_productos(): self.tabla_productos.insert('', 'end', values=f)
        self.refrescar_combobox()

    def nuevo_producto(self):
        top = Toplevel(self); top.title('Nuevo producto')
        ttk.Label(top, text='Nombre').grid(row=0, column=0, padx=5, pady=5, sticky='w')
        ttk.Label(top, text='Existencia').grid(row=1, column=0, padx=5, pady=5, sticky='w')
        ttk.Label(top, text='Proveedor').grid(row=2, column=0, padx=5, pady=5, sticky='w')
        ttk.Label(top, text='Precio').grid(row=3, column=0, padx=5, pady=5, sticky='w')
        e2 = ttk.Entry(top); e3 = ttk.Entry(top); e4 = ttk.Entry(top); e5 = ttk.Entry(top)
        e2.grid(row=0, column=1); e3.grid(row=1, column=1); e4.grid(row=2, column=1); e5.grid(row=3, column=1)
        def guardar():
            ok, msg = crear_producto(e2.get().strip(), e3.get().strip(), e4.get().strip(), e5.get().strip())
            messagebox.showinfo('Info', msg); 
            if ok: top.destroy(); self.refrescar_productos()
        ttk.Button(top, text='Guardar', command=guardar).grid(row=4, column=0, columnspan=2, pady=8)

    def editar_producto(self):
        sel = self.tabla_productos.selection()
        if not sel: messagebox.showwarning('Atención','Seleccione un producto'); return
        vals = self.tabla_productos.item(sel[0])['values']  # codigo, nombre, existencia, proveedor, precio
        top = Toplevel(self); top.title('Editar producto')
        ttk.Label(top, text=f'Código: {vals[0]}').grid(row=0, column=0, padx=5, pady=5, sticky='w', columnspan=2)
        ttk.Label(top, text='Nombre').grid(row=1, column=0, padx=5, pady=5, sticky='w')
        ttk.Label(top, text='Existencia').grid(row=2, column=0, padx=5, pady=5, sticky='w')
        ttk.Label(top, text='Proveedor').grid(row=3, column=0, padx=5, pady=5, sticky='w')
        ttk.Label(top, text='Precio').grid(row=4, column=0, padx=5, pady=5, sticky='w')
        e2 = ttk.Entry(top); e3 = ttk.Entry(top); e4 = ttk.Entry(top); e5 = ttk.Entry(top)
        e2.insert(0, vals[1]); e3.insert(0, vals[2]); e4.insert(0, vals[3]); e5.insert(0, vals[4])
        e2.grid(row=1, column=1); e3.grid(row=2, column=1); e4.grid(row=3, column=1); e5.grid(row=4, column=1)
        def guardar():
            ok, msg = actualizar_producto(vals[0], e2.get().strip(), e3.get().strip(), e4.get().strip(), e5.get().strip())
            messagebox.showinfo('Info', msg); 
            if ok: top.destroy(); self.refrescar_productos()
        ttk.Button(top, text='Guardar', command=guardar).grid(row=5, column=0, columnspan=2, pady=8)

    def cambiar_existencia(self):
        sel = self.tabla_productos.selection()
        if not sel: messagebox.showwarning('Atención','Seleccione un producto'); return
        vals = self.tabla_productos.item(sel[0])['values']
        top = Toplevel(self); top.title('Editar existencia')
        ttk.Label(top, text=f'Producto: {vals[1]}').grid(row=0, column=0, padx=5, pady=5, sticky='w', columnspan=2)
        ttk.Label(top, text='Nueva existencia').grid(row=1, column=0, padx=5, pady=5, sticky='w')
        e = ttk.Entry(top); e.insert(0, vals[2]); e.grid(row=1, column=1)
        def guardar():
            ok, msg = editar_existencia(vals[0], e.get().strip())
            messagebox.showinfo('Info', msg); 
            if ok: top.destroy(); self.refrescar_productos()
        ttk.Button(top, text='Guardar', command=guardar).grid(row=2, column=0, columnspan=2, pady=8)

    def eliminar_producto(self):
        sel = self.tabla_productos.selection()
        if not sel: messagebox.showwarning('Atención','Seleccione un producto'); return
        vals = self.tabla_productos.item(sel[0])['values']
        if messagebox.askyesno('Confirmar', f'¿Eliminar {vals[1]}?'):
            ok, msg = eliminar_producto(vals[0]); messagebox.showinfo('Info', msg)
            if ok: self.refrescar_productos()

    # --- Clientes (acciones y utilidades)
    def ordenar_tabla(self, tv, col, descendente):
        datos = [(tv.set(k, col), k) for k in tv.get_children('')]
        try: datos.sort(key=lambda t: float(t[0]), reverse=descendente)
        except: datos.sort(key=lambda t: t[0], reverse=descendente)
        for i, (val, k) in enumerate(datos): tv.move(k, '', i)
        tv.heading(col, command=lambda: self.ordenar_tabla(tv, col, not descendente))

    def filtrar_clientes(self):
        patron = self.ent_buscar_cliente.get().lower().strip()
        for i in self.tabla_clientes.get_children(): self.tabla_clientes.delete(i)
        filas = listar_clientes()
        if not filas: self.lbl_sin_clientes.pack(pady=10)
        else: self.lbl_sin_clientes.pack_forget()
        for fila in filas:
            if patron in ' '.join([str(x or '') for x in fila]).lower():
                self.tabla_clientes.insert('', 'end', values=fila)

    def exportar_clientes_ui(self):
        ruta = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel','*.xlsx')], initialfile='Clientes.xlsx')
        if not ruta: return
        exportar_clientes(ruta); messagebox.showinfo('Listo','Clientes exportados')

    def importar_clientes_ui(self):
        ruta = filedialog.askopenfilename(filetypes=[('Excel','*.xlsx')])
        if not ruta: return
        ok, msg = importar_clientes(ruta); messagebox.showinfo('Resultado', msg if ok else f'Error: {msg}')
        self.refrescar_clientes()

    def refrescar_clientes(self):
        for i in self.tabla_clientes.get_children(): self.tabla_clientes.delete(i)
        filas = listar_clientes()
        if not filas: self.lbl_sin_clientes.pack(pady=10)
        else: self.lbl_sin_clientes.pack_forget()
        for fila in filas: self.tabla_clientes.insert('', 'end', values=fila)
        self.refrescar_combobox()

    def nuevo_cliente(self):
        top = Toplevel(self); top.title('Nuevo cliente')
        ttk.Label(top, text='Nombre').grid(row=0, column=0, padx=5, pady=5, sticky='w')
        ttk.Label(top, text='Dirección').grid(row=1, column=0, padx=5, pady=5, sticky='w')
        ttk.Label(top, text='Teléfono').grid(row=2, column=0, padx=5, pady=5, sticky='w')
        ttk.Label(top, text='Correo').grid(row=3, column=0, padx=5, pady=5, sticky='w')
        e2 = ttk.Entry(top); e3 = ttk.Entry(top); e4 = ttk.Entry(top); e5 = ttk.Entry(top)
        e2.grid(row=0, column=1); e3.grid(row=1, column=1); e4.grid(row=2, column=1); e5.grid(row=3, column=1)
        def guardar():
            ok, msg = crear_cliente(e2.get().strip(), e3.get().strip(), e4.get().strip(), e5.get().strip())
            messagebox.showinfo('Info', msg)
            if ok: top.destroy(); self.refrescar_clientes()
        ttk.Button(top, text='Guardar', command=guardar).grid(row=4, column=0, columnspan=2, pady=8)

    def editar_cliente(self):
        sel = self.tabla_clientes.selection()
        if not sel: messagebox.showwarning('Atención','Seleccione un cliente'); return
        vals = self.tabla_clientes.item(sel[0])['values']
        top = Toplevel(self); top.title('Editar cliente')
        ttk.Label(top, text=f'Código: {vals[0]}').grid(row=0, column=0, padx=5, pady=5, sticky='w', columnspan=2)
        ttk.Label(top, text='Nombre').grid(row=1, column=0, padx=5, pady=5, sticky='w')
        ttk.Label(top, text='Dirección').grid(row=2, column=0, padx=5, pady=5, sticky='w')
        ttk.Label(top, text='Teléfono').grid(row=3, column=0, padx=5, pady=5, sticky='w')
        ttk.Label(top, text='Correo').grid(row=4, column=0, padx=5, pady=5, sticky='w')
        e2 = ttk.Entry(top); e3 = ttk.Entry(top); e4 = ttk.Entry(top); e5 = ttk.Entry(top)
        e2.insert(0, vals[1]); e3.insert(0, vals[2]); e4.insert(0, vals[3]); e5.insert(0, vals[4])
        e2.grid(row=1, column=1); e3.grid(row=2, column=1); e4.grid(row=3, column=1); e5.grid(row=4, column=1)
        def guardar():
            ok, msg = editar_cliente(vals[0], e2.get().strip(), e3.get().strip(), e4.get().strip(), e5.get().strip())
            messagebox.showinfo('Info', msg)
            if ok: top.destroy(); self.refrescar_clientes()
        ttk.Button(top, text='Guardar', command=guardar).grid(row=5, column=0, columnspan=2, pady=8)

    def eliminar_cliente(self):
        sel = self.tabla_clientes.selection()
        if not sel: messagebox.showwarning('Atención','Seleccione un cliente'); return
        vals = self.tabla_clientes.item(sel[0])['values']
        if messagebox.askyesno('Confirmar', f'¿Eliminar {vals[1]}?'):
            ok, msg = eliminar_cliente(vals[0]); messagebox.showinfo('Info', msg)
            if ok: self.refrescar_clientes()
    def limpiar_detalle_venta(self):
        # Coloca placeholders cuando no haya ventas
        if not hasattr(self, 'v_id'):
            return
        self.v_id.set('--'); self.v_prod.set('--'); self.v_cli.set('--')
        self.v_cant.set('--'); self.v_total.set('--'); self.v_fecha.set('--'); self.v_estado.set('--')

    def mostrar_venta_inferior(self, venta):
        """venta: [id, codigo_producto, codigo_cliente, cantidad, total, fecha, estado]"""
        if not hasattr(self, 'v_id'):
            return
        if not venta:
            self.limpiar_detalle_venta(); return
        prod = obtener_producto(venta[1])
        cli  = obtener_cliente(venta[2])
        self.v_id.set(str(venta[0]))
        self.v_prod.set(f"{venta[1]} - {prod['nombre'] if prod else 'N/D'}")
        self.v_cli.set(f"{venta[2]} - {cli['nombre'] if cli else 'N/D'}")
        self.v_cant.set(str(venta[3]))
        self.v_total.set(f"Q{round(float(venta[4]), 2)}")
        self.v_fecha.set(str(venta[5]))
        self.v_estado.set(str(venta[6]))

    def actualizar_detalle_venta_desde_ultima(self):
        ventas = listar_ventas()
        if ventas:
            self.mostrar_venta_inferior(ventas[-1])
        else:
            self.limpiar_detalle_venta()


    # --- Ventas y reportes ---
    def refrescar_combobox(self):
        prods = listar_productos(); clis = listar_clientes()
        if hasattr(self, 'cb_producto'): self.cb_producto['values'] = [f"{p[0]} - {p[1]}" for p in prods]
        if hasattr(self, 'cb_cliente'): self.cb_cliente['values'] = [f"{c[0]} - {c[1]}" for c in clis]

    def calcular_total(self):
        if not self.cb_producto.get() or not self.ent_cantidad.get():
            messagebox.showwarning('Atención','Seleccione producto y cantidad'); return
        cod = self.cb_producto.get().split(' - ')[0]; prod = obtener_producto(cod)
        try: cant = float(self.ent_cantidad.get())
        except: messagebox.showwarning('Atención','La cantidad debe ser numérica'); return
        if cant <= 0: messagebox.showwarning('Atención','La cantidad debe ser mayor que cero'); return
        if prod and cant > float(prod['existencia']):
            messagebox.showwarning('Atención', f'Existencia insuficiente (disp: {prod["existencia"]})'); return
        total = float(prod['precio']) * cant if prod else 0; self.lbl_total.config(text=f'Total: Q{round(total,2)}')

    def crear_venta_ui(self):
        if not self.cb_producto.get() or not self.cb_cliente.get() or not self.ent_cantidad.get():
            messagebox.showwarning('Atención','Complete los datos'); return
        codp = self.cb_producto.get().split(' - ')[0]; codc = self.cb_cliente.get().split(' - ')[0]
        ok, msg = crear_venta(codp, codc, self.ent_cantidad.get().strip()); messagebox.showinfo('Info', msg)
        if ok:
            if hasattr(self, 'ent_buscar_venta'):
                self.ent_buscar_venta.delete(0, END)
            self.refrescar_ventas(); self.refrescar_productos()
            self.ent_cantidad.delete(0, END); self.lbl_total.config(text='Total: Q0.00')
            self.actualizar_detalle_venta_desde_ultima()

    def anular_venta_ui(self):
        sel = self.tabla_ventas.selection()
        if not sel: messagebox.showwarning('Atención','Seleccione una venta'); return
        vals = self.tabla_ventas.item(sel[0])['values']
        if messagebox.askyesno('Confirmar', f'¿Anular {vals[0]}?'):
            ok, msg = anular_venta(vals[0]); messagebox.showinfo('Info', msg)
            if ok: self.refrescar_ventas(); self.refrescar_productos()
            self.actualizar_detalle_venta_desde_ultima()

    def refrescar_ventas(self):
        for i in self.tabla_ventas.get_children(): self.tabla_ventas.delete(i)
        for fila in listar_ventas(): self.tabla_ventas.insert('', 'end', values=fila)

    def filtrar_ventas(self):
        patron = self.ent_buscar_venta.get().lower().strip()
        for i in self.tabla_ventas.get_children(): self.tabla_ventas.delete(i)
        for fila in listar_ventas():
            if patron in ' '.join([str(x or '') for x in fila]).lower():
                self.tabla_ventas.insert('', 'end', values=fila)

    def reporte_cliente(self):
        ruta = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel','*.xlsx')], initialfile='Reporte_Ventas_por_Cliente.xlsx')
        if not ruta: return
        generar_reporte_por_cliente(ruta); self.archivo_rep_cliente.set(ruta); messagebox.showinfo('Listo','Reporte generado')

    def reporte_producto(self):
        ruta = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel','*.xlsx')], initialfile='Reporte_Ventas_por_Producto.xlsx')
        if not ruta: return
        generar_reporte_por_producto(ruta); self.archivo_rep_producto.set(ruta); messagebox.showinfo('Listo','Reporte generado')

    def enviar_reportes(self):
        correo = getattr(self, 'ent_correo', None).get().strip()
        clave = getattr(self, 'ent_pass', None).get().strip()
        destino = getattr(self, 'ent_destino', None).get().strip()
        adjuntos = []
        if getattr(self, 'archivo_rep_cliente', None) and self.archivo_rep_cliente.get(): adjuntos.append(self.archivo_rep_cliente.get())
        if getattr(self, 'archivo_rep_producto', None) and self.archivo_rep_producto.get(): adjuntos.append(self.archivo_rep_producto.get())
        if not correo or not clave or not destino or not adjuntos:
            messagebox.showwarning('Atención','Complete correo, contraseña, destino y genere al menos un reporte'); return
        try:
            msg = EmailMessage(); msg['Subject'] = 'Reportes de Ventas'; msg['From'] = correo; msg['To'] = destino
            msg.set_content('Se adjuntan los reportes solicitados.')
            for ruta in adjuntos:
                with open(ruta, 'rb') as f: datos = f.read()
                msg.add_attachment(datos, maintype='application', subtype='octet-stream', filename=os.path.basename(ruta))
            contexto = ssl.create_default_context()
            with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=contexto) as server:
                server.login(correo, clave); server.send_message(msg)
            self.estado_correo.set('Enviado'); messagebox.showinfo('Listo','Correo enviado')
        except Exception as e:
            self.estado_correo.set('Error'); messagebox.showerror('Error', str(e))

if __name__ == '__main__':
    asegurar_excel()
    app = App()
    app.mainloop()